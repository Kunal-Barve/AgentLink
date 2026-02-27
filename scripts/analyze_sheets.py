import openpyxl
from collections import defaultdict

SUBSCRIBED_PATH = r'd:\Work\Upwork\MakeDropBox_integration\CodeBase\Make-Integration\app\assets\google_sheets\Agents Subscribed.xlsx'
FEATURED_PATH = r'd:\Work\Upwork\MakeDropBox_integration\CodeBase\Make-Integration\app\assets\google_sheets\Featured Agent Controls.xlsx'

TARGETS = ["Sam Thompson", "Rodney Goodwin", "Aman Singh", "Aaron Boud"]

wb2 = openpyxl.load_workbook(FEATURED_PATH, data_only=True)
ws2 = wb2[wb2.sheetnames[0]]
headers2 = [cell.value for cell in ws2[1]]

commission_cols = [h for h in headers2 if h and (
    "Commission" in str(h) or "Marketing" in str(h) or "Discount" in str(h)
)]

# Print compact: row#, suburb, state, then each commission/marketing on one line
for target in TARGETS:
    print(f"\n{'='*70}")
    print(f"AGENT: {target}")
    print(f"{'='*70}")
    rows_found = []
    for row_idx, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = dict(zip(headers2, row))
        name = row_dict.get("Name") or ""
        if target.lower() in name.lower():
            rows_found.append((row_idx, row_dict))

    print(f"Total rows found: {len(rows_found)}")
    print(f"{'Row':<6} {'Suburb':<25} {'State':<6} {'<500k C':<14} {'500k-1m C':<14} {'1m-1.5m C':<14} {'1.5m-2m C':<14} {'<500k M':<14} {'1.5m-2m M':<14}")
    print("-"*120)
    for row_idx, rd in rows_found:
        print(
            f"{row_idx:<6} {str(rd.get('Suburb') or ''):<25} {str(rd.get('State') or ''):<6} "
            f"{str(rd.get('Less than $500k Commission') or ''):<14} "
            f"{str(rd.get('$500k-$1m Commission') or ''):<14} "
            f"{str(rd.get('$1m-$1.5m Commission') or ''):<14} "
            f"{str(rd.get('$1.5m-$2m Commission') or ''):<14} "
            f"{str(rd.get('Less than $500k Marketing') or ''):<14} "
            f"{str(rd.get('$1.5m-$2m Marketing') or ''):<14}"
        )

    # Check if any commission/marketing values differ across the rows
    print("\nDifferences check:")
    any_diff = False
    for col in commission_cols:
        values = set(str(rd.get(col) or "").strip() for _, rd in rows_found)
        if len(values) > 1:
            print(f"  *** DIFFERENT '{col}': {values}")
            any_diff = True
    if not any_diff:
        print("  ALL commission/marketing values are IDENTICAL across all rows.")

import sys; sys.exit(0)

# ============================================================
# 1. AGENTS SUBSCRIBED - Sheet1
# ============================================================
wb1 = openpyxl.load_workbook(SUBSCRIBED_PATH, data_only=True)
ws1 = wb1['Sheet1']
headers1 = [cell.value for cell in ws1[1]]
print("=== AGENTS SUBSCRIBED - Sheet1 Headers ===")
print(headers1)
print(f"Total rows (incl header): {ws1.max_row}")

agent_suburbs = defaultdict(list)
for row in ws1.iter_rows(min_row=2, values_only=True):
    name = row[0]
    state = row[3]
    suburb = row[4]
    sub_type = row[7]
    if name and suburb:
        agent_suburbs[name].append({"suburb": suburb, "state": state, "type": sub_type})

multi = {k: v for k, v in agent_suburbs.items() if len(v) > 1}
print(f"\nTotal unique agents: {len(agent_suburbs)}")
print(f"Agents subscribed to >1 suburb: {len(multi)}")
print("\n=== AGENTS WITH MULTIPLE SUBURBS (sorted by count) ===")
for agent, subs in sorted(multi.items(), key=lambda x: len(x[1]), reverse=True):
    suburbs_list = [(s["suburb"], s["state"]) for s in subs]
    print(f"  {agent} ({len(subs)}): {suburbs_list}")

# ============================================================
# 2. FEATURED AGENT CONTROLS - Sheet1
# ============================================================
wb2 = openpyxl.load_workbook(FEATURED_PATH, data_only=True)
print(f"\n=== FEATURED AGENT CONTROLS Sheet Names: {wb2.sheetnames} ===")

ws2 = wb2[wb2.sheetnames[0]]
headers2 = [cell.value for cell in ws2[1]]
print(f"\n=== FEATURED AGENT CONTROLS - Sheet1 Headers ===")
print(headers2)
print(f"Total rows (incl header): {ws2.max_row}")

commission_cols = [h for h in headers2 if h and ("Commission" in str(h) or "Marketing" in str(h) or "Discount" in str(h))]
print(f"\nCommission/Marketing/Discount columns ({len(commission_cols)}): {commission_cols}")

featured_agents = defaultdict(list)
for row in ws2.iter_rows(min_row=2, values_only=True):
    row_dict = dict(zip(headers2, row))
    name = row_dict.get("Name")
    if name:
        featured_agents[name].append(row_dict)

multi_featured = {k: v for k, v in featured_agents.items() if len(v) > 1}
print(f"\nAgents with >1 row in Featured Controls: {len(multi_featured)}")

# ============================================================
# 3. KEY QUESTION: Do commission/marketing/discount values differ per suburb?
# ============================================================
print("\n=== DO COMMISSION RATES DIFFER ACROSS SUBURBS FOR THE SAME AGENT? ===")
agents_with_diff = []
for agent, rows in featured_agents.items():
    if len(rows) > 1:
        diffs = []
        for col in commission_cols:
            values = set(str(r.get(col, "") or "").strip() for r in rows)
            if len(values) > 1:
                diffs.append(f"{col}: {values}")
        if diffs:
            agents_with_diff.append((agent, diffs))
            print(f"  DIFFERENT -> {agent}:")
            for d in diffs:
                print(f"    {d}")

if not agents_with_diff:
    print("  CONFIRMED: All commission/marketing/discount values are IDENTICAL across suburbs for every agent.")
    print("  Your assumption is CORRECT - only Suburb and State change, rates stay the same.")

# ============================================================
# 4. FRAZER YULE DEEP DIVE
# ============================================================
print("\n=== FRAZER YULE in Agents Subscribed ===")
fy_subscribed = []
for row in ws1.iter_rows(min_row=2, values_only=True):
    if row[0] and "Frazer" in str(row[0]):
        fy_subscribed.append(dict(zip(headers1, row)))
print(f"  Total rows: {len(fy_subscribed)}")
for r in fy_subscribed:
    print(f"  Suburb={r.get('Suburb')}, State={r.get('State')}, Type={r.get('Subscription Type')}")

print("\n=== FRAZER YULE in Featured Agent Controls ===")
fy_featured = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    row_dict = dict(zip(headers2, row))
    if row_dict.get("Name") and "Frazer" in str(row_dict.get("Name")):
        fy_featured.append(row_dict)
print(f"  Total rows: {len(fy_featured)}")
for r in fy_featured:
    print(f"  Suburb={r.get('Suburb')}, State={r.get('State')}, $1.5m-$2m Commission={r.get('$1.5m-$2m Commission')}, $1.5m-$2m Marketing={r.get('$1.5m-$2m Marketing')}")
