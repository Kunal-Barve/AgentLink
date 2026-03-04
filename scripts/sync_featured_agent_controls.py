"""
Sync Featured Agent Controls.xlsx Sheet1 -> Supabase featured_agent_controls table.

Clears existing rows then re-inserts all rows from the xlsx.

Run:
    python scripts/sync_featured_agent_controls.py

Prerequisites:
    pip install openpyxl supabase python-dotenv
    SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY in .env
"""

import os
import sys
from pathlib import Path
import openpyxl
from supabase import create_client, Client
from dotenv import load_dotenv

load_dotenv()

XLSX_PATH = Path(__file__).parent.parent / "app" / "assets" / "google_sheets" / "Featured Agent Controls.xlsx"
TABLE_NAME = "featured_agent_controls"
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
BATCH_SIZE = 200

# xlsx column name -> Supabase column name
# Featured Agent Controls has 27 columns: Name, Suburb, State + 12 commission + 12 marketing
COLUMN_MAP = {
    "Name":                         "name",
    "Suburb":                       "suburb",
    "State":                        "state",
    "Less than $500k Commission":   "less_than_500k_commission",
    "$500k-$1m Commission":         "500k_1m_commission",
    "$1m-$1.5m Commission":         "1m_1_5m_commission",
    "$1.5m-$2m Commission":         "1_5m_2m_commission",
    "$2m-$2.5m Commission":         "2m_2_5m_commission",
    "$2.5m-$3m Commission":         "2_5m_3m_commission",
    "$3-$3.5m Commission":          "3_3_5m_commission",
    "$3.5m-$4m Commission":         "3_5m_4m_commission",
    "$4m-$6m Commission":           "4m_6m_commission",
    "$6m-$8m Commission":           "6m_8m_commission",
    "$8m-$10m Commission":          "8m_10m_commission",
    "$10m+ Commission":             "10m_commission",
    "Less than $500k Marketing":    "less_than_500k_marketing",
    "$500k-$1m Marketing":          "500k_1m_marketing",
    "$1m-$1.5m Marketing":          "1m_1_5m_marketing",
    "$1.5m-$2m Marketing":          "1_5m_2m_marketing",
    "$2m-$2.5m Marketing":          "2m_2_5m_marketing",
    "$2.5m-$3m Marketing":          "2_5m_3m_marketing",
    "$3m-$3.5m Marketing":          "3m_3_5m_marketing",
    "$3.5m-$4m Marketing":          "3_5m_4m_marketing",
    "$4m-$6m Marketing":            "4m_6m_marketing",
    "$6m-$8m Marketing":            "6m_8m_marketing",
    "$8m-$10m Marketing":           "8m_10m_marketing",
    "$10m+ Marketing":              "10m_marketing",
}


def clean_value(val):
    if val is None or val == "":
        return None
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip() if str(val).strip() != "" else None


def load_xlsx(path: Path) -> list[dict]:
    print(f"Loading: {path}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) + [None] * (len(headers) - len(row))
        raw = dict(zip(headers, row))

        if not raw.get("Name"):
            continue

        record = {}
        for xlsx_col, db_col in COLUMN_MAP.items():
            record[db_col] = clean_value(raw.get(xlsx_col))

        rows.append(record)

    wb.close()
    print(f"Loaded {len(rows)} rows from xlsx")
    return rows


def sync(supabase: Client, rows: list[dict]):
    print(f"Clearing existing rows from '{TABLE_NAME}'...")
    supabase.table(TABLE_NAME).delete().neq("id", 0).execute()

    total = len(rows)
    inserted = 0
    for i in range(0, total, BATCH_SIZE):
        batch = rows[i: i + BATCH_SIZE]
        supabase.table(TABLE_NAME).insert(batch).execute()
        inserted += len(batch)
        print(f"  Inserted {inserted}/{total} rows ({inserted/total*100:.0f}%)")

    print(f"\n✅ Done — {inserted} rows synced to '{TABLE_NAME}'")


def main():
    if not SUPABASE_URL or not SUPABASE_KEY:
        print("ERROR: SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set in .env")
        sys.exit(1)
    if not XLSX_PATH.exists():
        print(f"ERROR: File not found: {XLSX_PATH}")
        sys.exit(1)

    print("=== Sync: Featured Agent Controls → Supabase ===\n")
    rows = load_xlsx(XLSX_PATH)

    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    sync(supabase, rows)


if __name__ == "__main__":
    main()
